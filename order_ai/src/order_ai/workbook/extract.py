from io import BytesIO

import openpyxl
import xlrd

from order_ai.models import Document, WorkbookSheet, Workbook


def load_workbook(document: Document) -> Workbook:
    """
    Load an Excel workbook into a simple in-memory representation.

    The rest of the pipeline should never depend directly on
    openpyxl/xlrd objects.
    """

    if document.binary is None:
        raise ValueError("Document has no binary data.")

    filename = (document.filename or "").lower()

    if filename.endswith(".xlsx"):
        return _load_xlsx(document.binary)

    if filename.endswith(".xls"):
        return _load_xls(document.binary)

    raise ValueError(f"Unsupported spreadsheet: {document.filename}")


def _load_xlsx(data: bytes) -> Workbook:
    workbook = openpyxl.load_workbook(
        BytesIO(data),
        data_only=True,
        read_only=True,
    )

    sheets: list[WorkbookSheet] = []

    for ws in workbook.worksheets:
        rows: list[list[str]] = []

        for row in ws.iter_rows(values_only=True):
            values = [
                "" if cell is None else str(cell).strip()
                for cell in row
            ]

            # Skip completely empty rows
            if any(values):
                rows.append(values)

        sheets.append(
            WorkbookSheet(
                name=ws.title,
                rows=rows,
            )
        )

    workbook.close()

    return Workbook(
        sheets=sheets,
    )


def _load_xls(data: bytes) -> _load_xlsx:
    workbook = xlrd.open_workbook(file_contents=data)

    sheets: list[WorkbookSheet] = []

    for ws in workbook.sheets():
        rows: list[list[str]] = []

        for r in range(ws.nrows):
            values = [
                "" if cell.value is None else str(cell.value).strip()
                for cell in ws.row(r)
            ]

            if any(values):
                rows.append(values)

        sheets.append(
            WorkbookSheet(
                name=ws.name,
                rows=rows,
            )
        )

    return Workbook(
        sheets=sheets,
    )
