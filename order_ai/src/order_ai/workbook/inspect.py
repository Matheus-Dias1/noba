from enum import StrEnum
from io import BytesIO

import openpyxl
import xlrd
from pydantic import BaseModel

from order_ai.models import EmailAttachment, AttachmentKind

class SheetInfo(BaseModel):
    name: str
    rows: int
    columns: int


class AttachmentInspection(BaseModel):
    filename: str
    mime_type: str
    size: int
    kind: AttachmentKind
    sheets: list[SheetInfo] = []


def inspect_attachment(att: EmailAttachment) -> AttachmentInspection:
    """
    Inspect an attachment and return basic metadata.

    This function never performs AI extraction. It only identifies the
    attachment type and gathers enough information for the next pipeline
    stage to decide how to process it.
    """

    inspection = AttachmentInspection(
        filename=att.filename,
        mime_type=att.mime_type,
        size=len(att.data),
        kind=AttachmentKind.UNKNOWN,
    )

    filename = att.filename.lower()
    mime = att.mime_type

    # ------------------------------------------------------------------
    # Excel (.xlsx)
    # ------------------------------------------------------------------

    if filename.endswith(".xlsx"):
        inspection.kind = AttachmentKind.SPREADSHEET

        workbook = openpyxl.load_workbook(
            BytesIO(att.data),
            read_only=True,
            data_only=True,
        )

        inspection.sheets = [
            SheetInfo(
                name=sheet.title,
                rows=sheet.max_row,
                columns=sheet.max_column,
            )
            for sheet in workbook.worksheets
        ]

        workbook.close()
        return inspection

    # ------------------------------------------------------------------
    # Legacy Excel (.xls)
    # ------------------------------------------------------------------

    if filename.endswith(".xls"):
        inspection.kind = AttachmentKind.SPREADSHEET

        workbook = xlrd.open_workbook(file_contents=att.data)

        inspection.sheets = [
            SheetInfo(
                name=sheet.name,
                rows=sheet.nrows,
                columns=sheet.ncols,
            )
            for sheet in workbook.sheets()
        ]

        return inspection

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    if mime == "application/pdf" or filename.endswith(".pdf"):
        inspection.kind = AttachmentKind.PDF
        return inspection

    # ------------------------------------------------------------------
    # Images
    # ------------------------------------------------------------------

    if mime.startswith("image/"):
        inspection.kind = AttachmentKind.IMAGE
        return inspection

    # ------------------------------------------------------------------
    # Word documents
    # ------------------------------------------------------------------

    if (
        filename.endswith(".doc")
        or filename.endswith(".docx")
        or mime
        == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    ):
        inspection.kind = AttachmentKind.DOCUMENT
        return inspection

    return inspection
